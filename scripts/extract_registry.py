#!/usr/bin/env python3
"""Extract the CRMDR canonical ID registry from the private digitization workbook.

The workbook ("Roman Martyrology LA IT EN with IDs.xlsx") contains the full
Latin, Italian and English texts of the eulogies of the Roman Martyrology
(editio altera 2004). Those texts are copyrighted and are NOT extracted here:
this script reads only the non-copyrightable structural metadata — the
canonical ID, calendar placement (month/day/entry), the asterisk marker, and
the country associated with the place of the elogium — and writes:

  data/martyrology_ids.json   machine-readable registry
  registry/MM-<month>.md      human-readable per-month tables

Usage:
  python3 extract_registry.py /path/to/"Roman Martyrology LA IT EN with IDs.xlsx" [repo_root]

Requires: openpyxl
"""

import json
import sys
from pathlib import Path

import openpyxl

MONTH_SHEETS = [
    "GENNAIO", "FEBBRAIO", "MARZO", "APRILE", "MAGGIO", "GIUGNO",
    "LUGLIO", "AGOSTO", "SETTEMBRE", "OTTOBRE", "NOVEMBRE", "DICEMBRE",
]
MONTH_NAMES_EN = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# IDs in the workbook that violate the canonicalization rules and are
# corrected on extraction. Surnames are not latinized unless an already
# well-known Latin form exists: "Miki" is a Japanese surname with no such form.
ID_CORRECTIONS = {
    "mr:0206-paulus-mikus-et-socii": "mr:0206-paulus-miki-et-socii",
    # The Latin editio altera 2004 places Bl. Marcantonio Durando at June 10
    # (entry 9*); the Italian (CEI) edition places him at December 10 (entry
    # 9*). The MMDD anchor follows the Latin print.
    "mr:1210-marcus-antonius-durando": "mr:0610-marcus-antonius-durando",
}

# Placement overrides: the digitized workbook records the Italian (CEI)
# placement, but the anchor edition (the Latin editio altera 2004 print)
# places the elogium on a different day. Maps corrected ID -> (month, day,
# entry, note). `entry` is None because the workbook has no row for the
# anchor-edition day.
PLACEMENT_OVERRIDES = {
    "mr:0610-marcus-antonius-durando": (
        6, 10, None,
        "Entry 9* at June 10 in the Latin editio altera 2004 print (verified "
        "on the page scan); the Italian (CEI) edition and the digitized "
        "workbook place the elogium at December 10 (entry 9*).",
    ),
}

# Entries present in the Italian (CEI) edition and the digitized workbook but
# absent from the Latin editio altera 2004 print (all verified on the page
# scans of both editions, July 2026).
ENTRY_NOTES = {
    "mr:0712-proclus-et-hilarion": (
        "Entry 1 at July 12 in the Italian (CEI) edition; absent from the "
        "Latin editio altera 2004 print, whose July 12 numbering begins at 2 "
        "with the gap left unrenumbered."
    ),
    "mr:0825-eusebius-et-socii": (
        "Entry 3 at August 25 in the Italian (CEI) edition; absent from the "
        "Latin editio altera 2004 print, where the day's numbered entries "
        "begin at 3 (Genesius) after the two unnumbered memorias."
    ),
    "mr:0709-maria-a-iesu-crucifixo-petkovic": (
        "Entry 11* at July 9 in the Italian (CEI) edition; absent from the "
        "Latin editio altera 2004 print, whose July 9 ends at entry 10*. "
        "Bl. Marija Petković was beatified on 6 June 2003."
    ),
}

# Asterisk overrides where the digitized workbook follows the Italian (CEI)
# edition but the anchor edition (the Latin editio altera 2004 print) differs.
# All were verified against the page scans of BOTH printed editions (2026-07),
# independently of the workbook: asterisk-presence claims via the OCR text
# layer with visual spot-checks, asterisk-absence claims each visually
# confirmed on the scan of the edition concerned.

# Entries asterisked in the Latin print but not in the CEI edition (nor in the
# workbook). Maps ID -> entry number as printed in the Latin editio altera.
LATIN_ASTERISKED = {
    "mr:0104-ferreolus": 4,
    "mr:0104-rigomerus": 5,
    "mr:0104-pharaildis": 7,
    "mr:0122-ladislaus-batthyany-strattmann": 15,
    "mr:0502-boleslaus-strzelecki": 12,
    "mr:0512-imelda-lambertini": 10,
    "mr:0524-servulus": 4,
    "mr:0611-bardo": 4,
    "mr:0624-gohardus": 7,
    "mr:0711-leontius": 5,
    "mr:0718-tarsicia-mackiv": 13,
    "mr:0730-godeleva": 6,
    "mr:0802-betharius": 8,
    "mr:0807-donatus-vesontione": 7,
    "mr:0909-franciscus-garate-aranguren": 10,
    "mr:1021-wendelinus": 8,
    "mr:1026-eata": 7,
    "mr:1103-libertinus": 3,
    "mr:1103-odrada": 10,
    "mr:1118-theofredus": 6,
    "mr:1119-eudo": 6,
    "mr:1215-marinus": 3,
    "mr:1230-egwinus": 7,
}

# Entries with a plain number in the Latin print that the CEI edition (and the
# workbook) marks with an asterisk. Maps ID -> entry number in the Latin print.
LATIN_PLAIN = {
    "mr:0323-rebecca-de-himlaya": 11,
    "mr:0812-iacobus": 11,
    "mr:0821-iosephus": 11,
    "mr:1201-domnolus": 5,
    "mr:1203-lucius": 5,
    "mr:1212-simon-phan": 11,
}

ASTERISK_OVERRIDES = {}
for _id, _n in LATIN_ASTERISKED.items():
    ASTERISK_OVERRIDES[_id] = (
        True,
        f"Asterisked entry ({_n}*) in the Latin editio altera 2004 print; "
        "the Italian (CEI) edition carries no asterisk.",
    )
for _id, _n in LATIN_PLAIN.items():
    ASTERISK_OVERRIDES[_id] = (
        False,
        f"Plain entry ({_n}., no asterisk) in the Latin editio altera 2004 "
        "print, visually verified on the page scan; the Italian (CEI) edition "
        "marks the entry with an asterisk.",
    )

# Entries present in the editio altera 2004 print but absent from the
# digitized "with IDs" workbook (see docs/canonicalization-report.md). Both
# are present in the parallel-texts workbook, whose reviewer comments document
# their numbering across editions.
PRINT_ONLY_ENTRIES = [
    {
        "id": "mr:0104-abrunculus",
        "month": 1,
        "day": 4,
        "entry": None,
        "asterisk": True,
        "country": "FR",
        "note": "Entry 2* in the Latin editio altera 2004 print; absent from "
                "the Italian (CEI) edition, from Mons. Barba's Word "
                "transcription, and from the digitized workbook.",
    },
    {
        "id": "mr:0104-emmanuel-gonzalez-garcia",
        "month": 1,
        "day": 4,
        "entry": None,
        "asterisk": True,
        "country": "ES",
        "note": "Numbered 12* in the Latin editio altera 2004 print, 11* in "
                "the Italian (CEI) edition and in Mons. Barba's Word "
                "transcription; absent from the digitized workbook. "
                "Bl. Manuel González García was canonized in 2016: status "
                "change with no ID change.",
    },
]

# The four leap-day elogia are printed twice (Feb 28 and Feb 29) and carry a
# single identity each, anchored at 0229. The Feb 29 placement is primary;
# the Feb 28 placement is recorded in `also_on`.
LEAP_DAY_PRIMARY = (2, 29)


def build_country_map(wb):
    iso = {}
    ws = wb["Paesi"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, _en, it = (list(row) + [None] * 3)[:3]
        if code and it:
            iso[it.strip()] = code.strip()
    return iso


def extract(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    iso = build_country_map(wb)
    rows = []
    for month_index, sheet in enumerate(MONTH_SHEETS, 1):
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=3, values_only=True):
            _mese, giorno, voce, asterisk, _it, _en, _la, paese, mr_id = (
                list(row) + [None] * 9
            )[:9]
            if mr_id is None:
                continue
            paese = str(paese).strip() if paese is not None else ""
            country = iso[paese] if paese else None
            mr_id = str(mr_id).strip()
            mr_id = ID_CORRECTIONS.get(mr_id, mr_id)
            row_out = {
                "id": mr_id,
                "month": month_index,
                "day": int(str(giorno)),
                "entry": int(str(voce)),
                "asterisk": asterisk == "*",
                "country": country,
            }
            if mr_id in ASTERISK_OVERRIDES:
                row_out["asterisk"], row_out["note"] = ASTERISK_OVERRIDES[mr_id]
            if mr_id in PLACEMENT_OVERRIDES:
                p_month, p_day, p_entry, p_note = PLACEMENT_OVERRIDES[mr_id]
                row_out.update(month=p_month, day=p_day, entry=p_entry)
                row_out["note"] = (row_out.get("note", "") + " " + p_note).strip()
            if mr_id in ENTRY_NOTES:
                row_out["note"] = (row_out.get("note", "") + " " + ENTRY_NOTES[mr_id]).strip()
            rows.append(row_out)

    # Merge duplicate IDs (the leap-day elogia): keep the Feb 29 placement as
    # primary and record the other placement in `also_on`.
    by_id = {}
    entries = []
    for row in rows:
        if row["id"] in by_id:
            first = by_id[row["id"]]
            primary, secondary = (
                (row, first)
                if (row["month"], row["day"]) == LEAP_DAY_PRIMARY
                else (first, row)
            )
            primary["also_on"] = [{
                "month": secondary["month"],
                "day": secondary["day"],
                "entry": secondary["entry"],
            }]
            if first is not primary:
                entries[entries.index(first)] = primary
            by_id[row["id"]] = primary
        else:
            by_id[row["id"]] = row
            entries.append(row)

    entries.extend(PRINT_ONLY_ENTRIES)
    entries.sort(key=lambda e: (e["month"], e["day"], e["entry"] is None, e["entry"] or 0))
    return entries


def write_json(entries, repo_root):
    out = {
        "$comment": "Canonical ID registry for the eulogies of the Roman Martyrology. "
                    "IDs are drafts pending committee review; see the README and "
                    "docs/canonicalization-report.md.",
        "anchor_edition": "martyrologium_romanum_editio_altera_2004",
        "id_scheme": "mr:MMDD-slug",
        "entry_count": len(entries),
        "entries": entries,
    }
    path = repo_root / "data" / "martyrology_ids.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
        f.write("\n")
    return path


def write_markdown(entries, repo_root):
    reg_dir = repo_root / "registry"
    reg_dir.mkdir(parents=True, exist_ok=True)
    for month_index, month_name in enumerate(MONTH_NAMES_EN, 1):
        month_entries = [e for e in entries if e["month"] == month_index]
        lines = [
            f"# {month_name}",
            "",
            f"{len(month_entries)} canonical IDs. "
            "`Entry` is the elogium's position within the day in the digitized workbook "
            "(editio altera 2004); `*` marks asterisked entries; `Country` is the "
            "ISO 3166-1 alpha-2 code of the modern country of the place of the elogium.",
            "",
            "| Day | Entry | ID | * | Country | Notes |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
        for e in month_entries:
            notes = []
            if e.get("also_on"):
                for a in e["also_on"]:
                    notes.append(f"also printed at {a['month']}/{a['day']} entry {a['entry']}")
            if e.get("note"):
                notes.append(e["note"])
            lines.append(
                f"| {e['day']} | {e['entry'] if e['entry'] is not None else '—'} "
                f"| `{e['id']}` | {'*' if e['asterisk'] else ''} "
                f"| {e['country'] or ''} | {' '.join(notes)} |"
            )
        path = reg_dir / f"{month_index:02d}-{month_name.lower()}.md"
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    workbook_path = sys.argv[1]
    repo_root = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).resolve().parent.parent
    entries = extract(workbook_path)
    path = write_json(entries, repo_root)
    write_markdown(entries, repo_root)
    ids = [e["id"] for e in entries]
    assert len(ids) == len(set(ids)), "duplicate IDs after merge"
    print(f"Wrote {len(entries)} entries to {path} and registry/*.md")


if __name__ == "__main__":
    main()
